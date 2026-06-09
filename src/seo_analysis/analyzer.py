"""On-page SEO analysis driven by an Excel workbook.

Behavior is preserved from the original SEO.py: read keywords (column A) and
URLs (column B) from the given worksheet, fetch each page, and fill columns
C-T with title/description/header/image/link/video/list analysis, colour-coding
cells. The only change is that the workbook path and sheet name are now
parameters (defaulting to "Test.xlsx" / "Sheet1") so it can be packaged.
"""

from datetime import datetime
from urllib.parse import urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill

RED = [255, 0, 0]
YELLOW = [255, 255, 0]
WHITE = [255, 255, 255]


def _fill(cell, rgb):
    color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
    cell.fill = PatternFill(
        fill_type="solid",
        start_color="FF" + color_string,
        end_color="FF" + color_string,
    )


def _change_count(sheet, index, keywords, col):
    count = 0
    for kw in keywords:
        if kw.lower() in sheet[col + index].value.lower():
            count += 1
    return count


def _change_color(sheet, index, count, col, lnth):
    if count == 0:
        _fill(sheet[col + index], RED)
    elif count > 0 and count < lnth:
        _fill(sheet[col + index], YELLOW)
    else:
        _fill(sheet[col + index], WHITE)


def _header_count(sheet, soup, index, keywords, tag, col, lnth):
    countall = 0
    countfull = 0
    countpart = 0

    for h in soup.findAll(tag):
        countall += 1
        counttemp = 0
        for kw in keywords:
            if kw.lower() in h.get_text().lower():
                counttemp += 1

        if counttemp == lnth:
            countfull += 1
        elif counttemp > 0 and counttemp < lnth:
            countpart += 1

    sheet[col + index] = str(countall) + " - " + str(countfull) + " - " + str(countpart)

    if countall == 0:
        _fill(sheet[col + index], RED)
    else:
        _fill(sheet[col + index], WHITE)


def analyze(filepath="Test.xlsx", sheet_name="Sheet1", timeout=None, user_agent=None):
    """Run the SEO analysis over the workbook at ``filepath``.

    This mirrors the original script exactly, including its console output and
    error handling, so existing usage (analysing ``Test.xlsx``) is unchanged.

    ``timeout`` (seconds) and ``user_agent`` are optional. When left as ``None``
    the HTTP request behaves identically to the original (no timeout, default
    requests User-Agent).
    """
    headers = {"User-Agent": user_agent} if user_agent else None

    try:

        start = datetime.now()
        print("\nprogram started at " + str(start))

        wb = openpyxl.load_workbook(filepath)
        sheet = wb[sheet_name]
        list_URL = []

        for cellObj in sheet["B"]:
            list_URL.append(cellObj.value)

        for idx, url in enumerate(list_URL[1:]):
            if idx != "" and idx != None and url != "" and url != None:
                print(
                    "\nprocessing ["
                    + str(idx + 1)
                    + "/"
                    + str(len(list_URL) - 1)
                    + "] --> "
                    + url
                )

                page = requests.get(url, timeout=timeout, headers=headers)
                soup = BeautifulSoup(page.text, "lxml")
                index = str(idx + 2)
                keywords = str(sheet["A" + index].value).split()
                lnth = len(keywords)

                sheet["C" + index].value = sheet["B" + index].value
                count = 0
                for kw in keywords:
                    if kw.lower() in sheet["B" + index].value.lower():
                        count += 1
                _change_color(sheet, index, count, "C", lnth)

                if soup.title == None:
                    sheet["D" + index] = "NO TITLE FOUND"
                    _fill(sheet["D" + index], RED)
                else:
                    if soup.title.string != None:
                        sheet["D" + index] = soup.title.string
                        count = 0
                        for kw in keywords:
                            if kw.lower() in soup.title.string.lower():
                                count += 1
                        _change_color(sheet, index, count, "D", lnth)
                    else:
                        sheet["D" + index] = "NO TITLE FOUND"
                        _fill(sheet["D" + index], RED)

                desc = soup.find(attrs={"name": "Description"})

                if desc == None:
                    desc = soup.find(attrs={"name": "description"})

                if desc != None:
                    if desc["content"] == None or len(desc["content"].strip()) == 0:
                        sheet["E" + index] = "NO DESCRIPTION FOUND"
                    else:
                        sheet["E" + index] = desc["content"]
                else:
                    sheet["E" + index] = "NO DESCRIPTION FOUND"

                count = _change_count(sheet, index, keywords, "E")
                _change_color(sheet, index, count, "E", lnth)

                count = 0
                for h1 in soup.findAll("h1"):
                    count += 1

                if count == 0:
                    sheet["F" + index] = "NO H1 FOUND"
                elif count == 1:
                    if soup.find("h1") != None:
                        if len(soup.find("h1").get_text()) == 0:
                            sheet["F" + index] = "NO H1 FOUND"
                        else:
                            sheet["F" + index] = soup.find("h1").get_text()
                    else:
                        sheet["F" + index] = "NO H1 FOUND"
                else:
                    sheet["F" + index] = (
                        "MEHR ALS 2 H1-ÜBERSCHRIFTEN (" + str(count) + " H1 FOUND)"
                    )
                    sheet["F" + index].font = Font(bold=True)

                count = _change_count(sheet, index, keywords, "F")
                _change_color(sheet, index, count, "F", lnth)

                _header_count(sheet, soup, index, keywords, "h2", "G", lnth)
                _header_count(sheet, soup, index, keywords, "h3", "H", lnth)
                _header_count(sheet, soup, index, keywords, "h4", "I", lnth)
                _header_count(sheet, soup, index, keywords, "h5", "J", lnth)
                _header_count(sheet, soup, index, keywords, "h6", "K", lnth)

                imflag = 0
                for im in soup.findAll("img"):
                    wd = im.get("width")
                    if wd == None:
                        wd = 0
                    else:
                        wd = wd.replace(";", "")
                        wd = wd.replace("p", "")
                        wd = wd.replace("x", "")

                    if int(wd) >= 300:
                        ht = im.get("height")
                        if ht == None:
                            ht = 0
                        else:
                            ht = ht.replace(";", "")
                            ht = ht.replace("p", "")
                            ht = ht.replace("x", "")

                        if int(ht) >= 300:
                            src = im.get("src")
                            if src == None or len(src.strip()) == 0:
                                src = "NO SRC FOUND"

                            alt = im.get("alt")
                            if alt == None or len(alt.strip()) == 0:
                                alt = "NO ALT FOUND"

                            title = im.get("title")
                            if title == None or len(title.strip()) == 0:
                                title = "NO IMAGE-TITLE FOUND"

                            sheet["L" + index] = src
                            sheet["M" + index] = alt
                            sheet["N" + index] = title
                            sheet["O" + index] = ht
                            sheet["P" + index] = wd

                            imflag = 1
                            break

                if not imflag:
                    sheet["L" + index] = "BILD FEHLT"
                    sheet["L" + index].font = Font(bold=True)
                    sheet["M" + index] = "BILD FEHLT"
                    sheet["M" + index].font = Font(bold=True)
                    sheet["N" + index] = "BILD FEHLT"
                    sheet["N" + index].font = Font(bold=True)

                count = _change_count(sheet, index, keywords, "L")
                _change_color(sheet, index, count, "L", lnth)

                count = _change_count(sheet, index, keywords, "M")
                _change_color(sheet, index, count, "M", lnth)

                count = _change_count(sheet, index, keywords, "N")
                _change_color(sheet, index, count, "N", lnth)

                extlink = 0
                intlink = 0

                for a in soup.findAll("a", attrs={"href": True}):
                    if (
                        len(a["href"].strip()) > 1
                        and a["href"][0] != "#"
                        and "javascript:" not in a["href"].strip()
                        and "mailto:" not in a["href"].strip()
                        and "tel:" not in a["href"].strip()
                    ):
                        if "http" in a["href"].strip() or "https" in a["href"].strip():
                            if (
                                urlparse(sheet["B" + index].value).netloc.lower()
                                in urlparse(a["href"].strip()).netloc.lower()
                            ):
                                intlink += 1
                            else:
                                extlink += 1
                        else:
                            intlink += 1

                sheet["Q" + index] = intlink
                sheet["R" + index] = extlink

                if intlink == 0:
                    _fill(sheet["Q" + index], RED)
                else:
                    _fill(sheet["Q" + index], WHITE)

                if extlink == 0:
                    _fill(sheet["R" + index], RED)
                else:
                    _fill(sheet["R" + index], WHITE)

                imflag = 0
                for ifr in soup.findAll("iframe", attrs={"src": True}):
                    if "youtube.com" in ifr["src"]:
                        imflag = 1
                        sheet["S" + index] = ifr["src"]
                        _fill(sheet["S" + index], WHITE)
                        break

                if not imflag:
                    sheet["S" + index] = "NO YOUTUBE VIDEO FOUND"
                    _fill(sheet["S" + index], RED)

                li = soup.findAll("li")
                sheet["T" + index] = len(li)
                if len(li) == 0:
                    _fill(sheet["T" + index], RED)
                else:
                    _fill(sheet["T" + index], WHITE)

            else:
                print(
                    "\nprocessing ["
                    + str(idx + 1)
                    + "/"
                    + str(len(list_URL) - 1)
                    + "] --> SKIPPING.. INVALID DOMAIN FOUND.."
                )

        wb.save(filepath)
        end = datetime.now()
        print("\nprogram finished at " + str(end))
        print(
            "\ntotal time taken is "
            + str((end - start).seconds)
            + "."
            + str((end - start).microseconds)
            + " seconds"
        )

    except FileNotFoundError:
        print("\ncould not find the file (" + filepath + "), please check path\n")

    except PermissionError:
        print("\nfile save failed, please close the file and run program again\n")

    except IOError:
        print(
            "\nwebsite ("
            + url
            + ") not found, exiting program, correct URL and run program again\n"
        )

    except (NameError, TypeError, RuntimeError, KeyError):
        print("\nsomething went wrong, exiting program\n")

    except Exception as e:
        print(e)
