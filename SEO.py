import re
from datetime import datetime
from urllib.parse import urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill


def changecount(keywords, col):
    count = 0
    for kw in keywords:
        if kw.lower() in sheet[col + index].value.lower():
            count += 1
    return count


def changecolor(count, col, lnth):
    if count == 0:
        rgb = [255, 0, 0]
        color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
        sheet[col + index].fill = PatternFill(
            fill_type="solid",
            start_color="FF" + color_string,
            end_color="FF" + color_string,
        )

    elif count > 0 and count < lnth:
        rgb = [255, 255, 0]
        color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
        sheet[col + index].fill = PatternFill(
            fill_type="solid",
            start_color="FF" + color_string,
            end_color="FF" + color_string,
        )

    else:
        rgb = [255, 255, 255]
        color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
        sheet[col + index].fill = PatternFill(
            fill_type="solid",
            start_color="FF" + color_string,
            end_color="FF" + color_string,
        )


def headercount(tag, col, lnth):
    countall = 0
    countfull = 0
    countpart = 0

    for h in soup.findAll(tag):
        countall += 1
        counttemp = 0
        for kw in keywords:
            if kw.lower() in h.get_text().lower():
                counttemp += 1

        if counttemp == lnth:
            countfull += 1
        elif counttemp > 0 and counttemp < lnth:
            countpart += 1

    sheet[col + index] = str(countall) + " - " + str(countfull) + " - " + str(countpart)

    if countall == 0:
        rgb = [255, 0, 0]
        color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
        sheet[col + index].fill = PatternFill(
            fill_type="solid",
            start_color="FF" + color_string,
            end_color="FF" + color_string,
        )
    else:
        rgb = [255, 255, 255]
        color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
        sheet[col + index].fill = PatternFill(
            fill_type="solid",
            start_color="FF" + color_string,
            end_color="FF" + color_string,
        )


try:

    start = datetime.now()
    print("\nprogram started at " + str(start))

    filepath = "Test.xlsx"
    wb = openpyxl.load_workbook(filepath)
    sheet = wb["Sheet1"]
    list_URL = []

    for cellObj in sheet["B"]:
        list_URL.append(cellObj.value)

    for idx, url in enumerate(list_URL[1:]):
        if idx != "" and idx != None and url != "" and url != None:
            print(
                "\nprocessing ["
                + str(idx + 1)
                + "/"
                + str(len(list_URL) - 1)
                + "] --> "
                + url
            )

            page = requests.get(url)
            soup = BeautifulSoup(page.text, "lxml")
            index = str(idx + 2)
            keywords = str(sheet["A" + index].value).split()
            lnth = len(keywords)

            sheet["C" + index].value = sheet["B" + index].value
            count = 0
            for kw in keywords:
                if kw.lower() in sheet["B" + index].value.lower():
                    count += 1
            changecolor(count, "C", lnth)

            if soup.title == None:
                sheet["D" + index] = "NO TITLE FOUND"
                rgb = [255, 0, 0]
                color_string = "".join(
                    [str(hex(i))[2:].upper().rjust(2, "0") for i in rgb]
                )
                sheet["D" + index].fill = PatternFill(
                    fill_type="solid",
                    start_color="FF" + color_string,
                    end_color="FF" + color_string,
                )

            else:
                if soup.title.string != None:
                    sheet["D" + index] = soup.title.string
                    count = 0
                    for kw in keywords:
                        if kw.lower() in soup.title.string.lower():
                            count += 1
                    changecolor(count, "D", lnth)

                else:
                    sheet["D" + index] = "NO TITLE FOUND"
                    rgb = [255, 0, 0]
                    color_string = "".join(
                        [str(hex(i))[2:].upper().rjust(2, "0") for i in rgb]
                    )
                    sheet["D" + index].fill = PatternFill(
                        fill_type="solid",
                        start_color="FF" + color_string,
                        end_color="FF" + color_string,
                    )

            desc = soup.find(attrs={"name": "Description"})

            if desc == None:
                desc = soup.find(attrs={"name": "description"})

            if desc != None:
                if desc["content"] == None or len(desc["content"].strip()) == 0:
                    sheet["E" + index] = "NO DESCRIPTION FOUND"
                else:
                    sheet["E" + index] = desc["content"]
            else:
                sheet["E" + index] = "NO DESCRIPTION FOUND"

            count = changecount(keywords, "E")
            changecolor(count, "E", lnth)
            count = 0
            for h1 in soup.findAll("h1"):
                count += 1

            if count == 0:
                sheet["F" + index] = "NO H1 FOUND"

            elif count == 1:
                if soup.find("h1") != None:
                    if len(soup.find("h1").get_text()) == 0:
                        sheet["F" + index] = "NO H1 FOUND"
                    else:
                        sheet["F" + index] = soup.find("h1").get_text()
                else:
                    sheet["F" + index] = "NO H1 FOUND"

            else:
                sheet["F" + index] = (
                    "MEHR ALS 2 H1-ÜBERSCHRIFTEN (" + str(count) + " H1 FOUND)"
                )
                sheet["F" + index].font = Font(bold=True)

            count = changecount(keywords, "F")
            changecolor(count, "F", lnth)

            headercount("h2", "G", lnth)
            headercount("h3", "H", lnth)
            headercount("h4", "I", lnth)
            headercount("h5", "J", lnth)
            headercount("h6", "K", lnth)

            imflag = 0
            for im in soup.findAll("img"):
                wd = im.get("width")
                if wd == None:
                    wd = 0
                else:
                    wd = wd.replace(";", "")
                    wd = wd.replace("p", "")
                    wd = wd.replace("x", "")

                if int(wd) >= 300:
                    ht = im.get("height")
                    if ht == None:
                        ht = 0
                    else:
                        ht = ht.replace(";", "")
                        ht = ht.replace("p", "")
                        ht = ht.replace("x", "")

                    if int(ht) >= 300:
                        src = im.get("src")
                        if src == None or len(src.strip()) == 0:
                            src = "NO SRC FOUND"

                        alt = im.get("alt")
                        if alt == None or len(alt.strip()) == 0:
                            alt = "NO ALT FOUND"

                        title = im.get("title")
                        if title == None or len(title.strip()) == 0:
                            title = "NO IMAGE-TITLE FOUND"

                        sheet["L" + index] = src
                        sheet["M" + index] = alt
                        sheet["N" + index] = title
                        sheet["O" + index] = ht
                        sheet["P" + index] = wd

                        imflag = 1
                        break

            if not imflag:
                sheet["L" + index] = "BILD FEHLT"
                sheet["L" + index].font = Font(bold=True)
                sheet["M" + index] = "BILD FEHLT"
                sheet["M" + index].font = Font(bold=True)
                sheet["N" + index] = "BILD FEHLT"
                sheet["N" + index].font = Font(bold=True)

            count = changecount(keywords, "L")
            changecolor(count, "L", lnth)

            count = changecount(keywords, "M")
            changecolor(count, "M", lnth)

            count = changecount(keywords, "N")
            changecolor(count, "N", lnth)

            extlink = 0
            intlink = 0

            for a in soup.findAll("a", attrs={"href": True}):
                if (
                    len(a["href"].strip()) > 1
                    and a["href"][0] != "#"
                    and "javascript:" not in a["href"].strip()
                    and "mailto:" not in a["href"].strip()
                    and "tel:" not in a["href"].strip()
                ):
                    if "http" in a["href"].strip() or "https" in a["href"].strip():
                        if (
                            urlparse(sheet["B" + index].value).netloc.lower()
                            in urlparse(a["href"].strip()).netloc.lower()
                        ):
                            intlink += 1
                        else:
                            extlink += 1
                    else:
                        intlink += 1

            sheet["Q" + index] = intlink
            sheet["R" + index] = extlink

            if intlink == 0:
                rgb = [255, 0, 0]
                color_string = "".join(
                    [str(hex(i))[2:].upper().rjust(2, "0") for i in rgb]
                )
                sheet["Q" + index].fill = PatternFill(
                    fill_type="solid",
                    start_color="FF" + color_string,
                    end_color="FF" + color_string,
                )
            else:
                rgb = [255, 255, 255]
                color_string = "".join(
                    [str(hex(i))[2:].upper().rjust(2, "0") for i in rgb]
                )
                sheet["Q" + index].fill = PatternFill(
                    fill_type="solid",
                    start_color="FF" + color_string,
                    end_color="FF" + color_string,
                )

            if extlink == 0:
                rgb = [255, 0, 0]
                color_string = "".join(
                    [str(hex(i))[2:].upper().rjust(2, "0") for i in rgb]
                )
                sheet["R" + index].fill = PatternFill(
                    fill_type="solid",
                    start_color="FF" + color_string,
                    end_color="FF" + color_string,
                )
            else:
                rgb = [255, 255, 255]
                color_string = "".join(
                    [str(hex(i))[2:].upper().rjust(2, "0") for i in rgb]
                )
                sheet["R" + index].fill = PatternFill(
                    fill_type="solid",
                    start_color="FF" + color_string,
                    end_color="FF" + color_string,
                )

            imflag = 0
            for ifr in soup.findAll("iframe", attrs={"src": True}):
                if "youtube.com" in ifr["src"]:
                    imflag = 1
                    sheet["S" + index] = ifr["src"]
                    rgb = [255, 255, 255]
                    color_string = "".join(
                        [str(hex(i))[2:].upper().rjust(2, "0") for i in rgb]
                    )
                    sheet["S" + index].fill = PatternFill(
                        fill_type="solid",
                        start_color="FF" + color_string,
                        end_color="FF" + color_string,
                    )
                    break

            if not imflag:
                sheet["S" + index] = "NO YOUTUBE VIDEO FOUND"
                rgb = [255, 0, 0]
                color_string = "".join(
                    [str(hex(i))[2:].upper().rjust(2, "0") for i in rgb]
                )
                sheet["S" + index].fill = PatternFill(
                    fill_type="solid",
                    start_color="FF" + color_string,
                    end_color="FF" + color_string,
                )

            li = soup.findAll("li")
            sheet["T" + index] = len(li)
            if len(li) == 0:
                rgb = [255, 0, 0]
                color_string = "".join(
                    [str(hex(i))[2:].upper().rjust(2, "0") for i in rgb]
                )
                sheet["T" + index].fill = PatternFill(
                    fill_type="solid",
                    start_color="FF" + color_string,
                    end_color="FF" + color_string,
                )
            else:
                rgb = [255, 255, 255]
                color_string = "".join(
                    [str(hex(i))[2:].upper().rjust(2, "0") for i in rgb]
                )
                sheet["T" + index].fill = PatternFill(
                    fill_type="solid",
                    start_color="FF" + color_string,
                    end_color="FF" + color_string,
                )

        else:
            print(
                "\nprocessing ["
                + str(idx + 1)
                + "/"
                + str(len(list_URL) - 1)
                + "] --> SKIPPING.. INVALID DOMAIN FOUND.."
            )

    wb.save(filepath)
    end = datetime.now()
    print("\nprogram finished at " + str(end))
    print(
        "\ntotal time taken is "
        + str((end - start).seconds)
        + "."
        + str((end - start).microseconds)
        + " seconds"
    )

except FileNotFoundError:
    print("\ncould not find the file (" + filepath + "), please check path\n")

except PermissionError:
    print("\nfile save failed, please close the file and run program again\n")

except IOError:
    print(
        "\nwebsite ("
        + url
        + ") not found, exiting program, correct URL and run program again\n"
    )

except (NameError, TypeError, RuntimeError, KeyError):
    print("\nsomething went wrong, exiting program\n")

except Exception as e:
    print(e)
